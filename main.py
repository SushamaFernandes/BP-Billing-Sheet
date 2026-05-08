import os
import re
import tkinter as tk
from tkinter import filedialog

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================================
# Domain configuration
# ============================================================================

RESOURCE_TO_TEAM = {
    "Punam Patil": "Dcab",
    "Paresh Damani": "Dcab",
    "Tejas Deshmukh": "Dcab",
    "Dattatray Awaghade": "Digistyle",
    "Swapnil Karekar": "Digistyle",
    "Prashant Bhayekar": "Product",
    "Dhawalshri Jadhav": "Product",
    "Anuja Redekar": "Product",
    "Abrar Shaha": "Product",
    "Ruchita Shetye": "Product",
    "Royston Rodrigues": "Product",
    "Sharad Kodag": "Product",
    "Vipin Verma": "Product",
    "Nishu Shah": "Product",
    "Chetan Adari": "Product",
    "Hrishikesh Dadhe": "Product",
    "Inderjeet Jethwani": "Product",
    "Priyanka Gupta": "Product",
    "Narayan Panigrahi": "Rebuying",
    "Sushama Fernandes": "Rebuying",
    "Dhiraj Pawar": "Rebuying",
    "Reshma Kute": "Rebuying",
    "Sagar Padwal": "Rebuying",
    "Shlok Patil": "Rebuying",
    "Mohd Waseem Shaikh": "Rebuying",
    "Ashwini Kanojia": "Rebuying",
    "Saurav Sharma": "Rebuying",
    "Shahid Shaikh": "Rebuying",
    "Pradnya Walkunde": "Rebuying",
    "Ajinkya Koparde": "Rebuying",
    "Kaustubh Chudji": "Rebuying",
    "Soham Kale": "AI",
    "Arun Kumar": "AI",
    "Rishi Misra": "AI",
    "Ajinkya Prabhu": "AI",
}

ISSUE_PREFIX_TO_MODULE = {
    "EKDM": "EKDM",
    "DPROD": "Product",
    "DSDEV": "Digistyle",
    "DCOM": "Commerce",
    "TAC": "TAC",
    "DART": "Artwork",
    "STYLEPOOL": "Stylepool",
    "LABELS": "Labelgenerator",
    "DQUAL": "d:pat",
    "DOC": "d:document",
    "DBUS": "BusinessPartner",
    "CPT": "CPT",
    "CAB": "A2DSS",
    "PRICE": "d:pricing",
    "DSDATA": "Commerce",
    "DSCM": "d:iwa",
    "DCAB": "Dcab",
    "DSEL": "d:select",
    "DCIS": "d:cision",
    "APL": "APL",
    "DQUAN": "d:quantity",
    "DINV": "d:invoice",
    "DROSI": "Dcab",
    "DMILE": "Milestonemaster",
    "DSIGN": "Product",
    "DMAR": "dmart",
    "ONPR": "Onlinepricing",
    "PRIPRI": "Printpricing",
    "MRSH": "Marshaller",
    "DPREP": "d:orderprep",
}

DSDATA_KEYWORD_TO_MODULE = {
    "commerce": "Commerce",
    "iwa": "d:iwa",
    "select": "d:select",
    "cision": "d:cision",
    "quantity": "d:quantity",
    "invoice": "d:invoice",
    "cab": "Dcab",
    "milestone": "Milestonemaster",
    "label": "Labelgenerator",
    "pricing": "d:pricing",
    "api": "API",
    "product": "Product",
    "artwork": "Artwork",
    "digistyle": "Digistyle",
    "tac": "TAC",
    "ekdm": "EKDM",
    "stylepool": "Stylepool",
    "businesspartner": "BusinessPartner",
    "document": "d:document",
    "pat": "d:pat",
    "a2dss": "A2DSS",
    "cpt": "CPT",
    "apl": "APL",
    "dmar": "dmart",
    "onpr": "Onlinepricing",
    "pripri": "Printpricing",
    "mrsh": "Marshaller",
    "dprep": "d:orderprep",
}

ANALYSIS_MEMBERS = {
    "Sushama Fernandes",
    "Prashant Bhayekar",
    "Dhawalshri Jadhav",
    "Punam Patil",
    "Dattatray Awaghade",
    "Nishu Shah",
    "Paresh Damani",
    "Sagar Padwal",
    "Swapnil Karekar",
    "Narayan Panigrahi",
    "Kaustubh Chudji",
}

TESTING_MEMBERS = {
    "Anuja Redekar",
    "Reshma Kute",
    "Pradnya Walkunde",
}

ALWAYS_BILLABLE_TASKS = {
    "coe weekly call",
    "functional testing of digistyle application",
    "interface call",
    "technical call",
    "showcase call",
    "sql coe dev dss",
    "technical call - bonprix",
    "telephonic call",
}

TEAM_RESTRICTED_BILLABLE_TASKS = {"framework call", "sprint planning", "scrum call"}
TEAM_RESTRICTED_BILLABLE_TEAMS = {"Digistyle", "Product"}

MEETING_TASKS = {
    "sprint-scrum-on call support",
    "coe weekly call",
    "telephonic call",
    "framework call",
    "scrum call",
    "technical call",
    "sprint planning",
}

DCAB_FORCED_MODULE_RESOURCES = {"Punam Patil", "Paresh Damani", "Tejas Deshmukh"}
DIGISTYLE_FORCED_MODULE_RESOURCES = {"Dattatray Awaghade", "Swapnil Karekar"}

AI_RESOURCES = {"Soham Kale", "Arun Kumar", "Rishi Misra", "Ajinkya Prabhu"}

AI_ROLE_MAP = {
    "Ajinkya Prabhu": "Senior Developer",
    "Rishi Misra": "Technical Architecture",
    "Arun Kumar": "Junior Developer",
    "Soham Kale": "Junior Developer",
}

AI_MODULE_RENAME = {"Gen AI": "ais-pim-genai-services"}

INPUT_REQUIRED_COLS = ["Entry Date", "Resource Name", "Task Name", "Actul Work(hrs)"]

HEADER_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=11, bold=False)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
SHEET_NAME = "Timesheet"


# ============================================================================
# Processing
# ============================================================================

class BillingProcessor:
    """Adds derived columns (Issue#, Teams, Module, Task Type, Billable, ...) to the input."""

    def process(self, df_in):
        missing = [c for c in INPUT_REQUIRED_COLS if c not in df_in.columns]
        if missing:
            raise ValueError(f"Missing column(s): {', '.join(missing)}")

        df = df_in.copy()
        df = self._filter_unwanted(df)
        df = self._format_date(df)
        df = self._extract_issue_number(df)
        df = self._clean_task_name(df)
        df["Teams"] = df.apply(self._infer_team, axis=1)
        df["Module"] = df.apply(self._infer_module, axis=1)
        df.loc[df["Resource Name"].isin(DCAB_FORCED_MODULE_RESOURCES), "Module"] = "Dcab"
        df.loc[df["Resource Name"].isin(DIGISTYLE_FORCED_MODULE_RESOURCES), "Module"] = "Digistyle"
        df["Task Type"] = df["Resource Name"].apply(self._infer_task_type)
        df["Mandays"] = ""  # populated as Excel formula at write time
        df["Billable"] = df.apply(self._infer_billable, axis=1)
        df["Module"] = df.apply(lambda row: self._fill_meeting_module(row, df), axis=1)
        df["Module"] = df["Module"].replace(np.nan, "", regex=True)
        return df

    @staticmethod
    def _filter_unwanted(df):
        if "Project Name" in df.columns:
            df = df[~df["Project Name"].str.contains(r"Training|NBT", case=False, na=False)]
        df = df[df["Resource Name"] != "Mary Stella"]
        return df

    @staticmethod
    def _format_date(df):
        df["Entry Date"] = pd.to_datetime(df["Entry Date"], errors="coerce").dt.strftime("%d-%m-%Y")
        return df

    @staticmethod
    def _extract_issue_number(df):
        df["Issue#"] = df["Task Name"].str.extract(
            r"([A-Za-z]+-\d+)", flags=re.IGNORECASE, expand=False
        )
        df["Issue#"] = df["Issue#"].where(df["Issue#"].notna(), pd.NA)
        return df

    @staticmethod
    def _clean_task_name(df):
        def clean(task_name):
            cleaned = re.sub(r"([A-Za-z]+-\d+)", "", str(task_name), flags=re.IGNORECASE).strip()
            cleaned = re.sub(r".*?\d*\s*-->(\s*:\s*|:\s*|\s*:|:|\s*)", "", cleaned)
            cleaned = re.sub(r"\s*:\s*", "", cleaned)
            return cleaned.strip()

        df["Task Name Clean"] = df["Task Name"].apply(clean)
        df["Task Name Out"] = np.where(
            df["Task Name Clean"] != "", df["Task Name Clean"], df["Task Name"]
        )
        return df

    @staticmethod
    def _infer_team(row):
        resource = row["Resource Name"]
        if resource in RESOURCE_TO_TEAM:
            return RESOURCE_TO_TEAM[resource]
        if "DJ" in str(row["Task Name"]):
            return "Product"
        return ""

    @staticmethod
    def _infer_module(row):
        task = str(row["Task Name"])
        issue_val = row.get("Issue#")
        task_lower = task.strip().lower()

        if (not pd.notna(issue_val)) or str(issue_val).strip() == "":
            if task_lower in MEETING_TASKS:
                return np.nan  # filled later from same-day modules

        if pd.notna(issue_val) and str(issue_val).strip() != "":
            prefix = str(issue_val).strip().split("-")[0].upper()
            if prefix == "DSDATA":
                low = task.lower()
                for keyword, module in DSDATA_KEYWORD_TO_MODULE.items():
                    if keyword in low:
                        return module
                return ""
            module = ISSUE_PREFIX_TO_MODULE.get(prefix, "")
            if module:
                return module

        if "API" in task:
            return "API"
        return ""

    @staticmethod
    def _infer_task_type(resource):
        if resource in ANALYSIS_MEMBERS:
            return "Analysis"
        if resource in TESTING_MEMBERS:
            return "Testing"
        return "Development"

    @staticmethod
    def _infer_billable(row):
        task_lower = str(row["Task Name"]).lower()
        task_stripped = task_lower.strip()

        if "leave" in task_lower:
            return "No"
        if task_stripped == "sprint-scrum-on call support":
            return "Yes"
        if task_stripped in ALWAYS_BILLABLE_TASKS:
            return "Yes"
        team = str(row.get("Teams", ""))
        if team in TEAM_RESTRICTED_BILLABLE_TEAMS and task_stripped in TEAM_RESTRICTED_BILLABLE_TASKS:
            return "Yes"
        if (pd.notna(row["Issue#"]) and str(row["Issue#"]).strip() != "") or "project management" in task_lower:
            return "Yes"
        if row["Task Type"] == "Analysis" and row["Teams"] == "Product":
            return "Yes"
        return "No"

    @staticmethod
    def _fill_meeting_module(row, df):
        task_name = str(row["Task Name"]).strip().lower()
        if task_name in {"sprint-scrum-on call support", "on leave"}:
            modules = df[
                (df["Resource Name"] == row["Resource Name"])
                & df["Module"].notna()
                & (df["Module"] != "")
            ]["Module"]
            if not modules.empty:
                most_common = modules.mode()
                if not most_common.empty:
                    return most_common.iloc[0]
        if pd.isna(row["Module"]) or row["Module"] == "":
            same_day = df[
                (df["Resource Name"] == row["Resource Name"])
                & (df["Entry Date"] == row["Entry Date"])
                & df["Module"].notna()
                & (df["Module"] != "")
            ]
            if not same_day.empty:
                most_common = same_day["Module"].mode()
                if not most_common.empty:
                    return most_common.iloc[0]
        return row["Module"]


# ============================================================================
# Output builders
# ============================================================================

class BillingOutputBuilder:
    """Standard billing output (Issue# column included), excluding AI team members."""

    COLUMNS = [
        "Entry Date", "Resource Name", "Issue#", "Task Name",
        "Actul Work(hrs)", "Teams", "Module", "Task Type", "Mandays", "Billable",
    ]
    MANDAYS_FORMULA = "=ROUNDUP({cell}/8,2)"
    COLUMN_WIDTHS = {
        "Entry Date": 14,
        "Resource Name": 20,
        "Issue#": 14,
        "Task Name": 60,
        "Actul Work(hrs)": 18,
        "Teams": 14,
        "Module": 18,
        "Task Type": 14,
        "Mandays": 12,
        "Billable": 12,
    }
    CENTERED_COLUMNS = ["Entry Date", "Issue#", "Actul Work(hrs)", "Teams", "Task Type", "Mandays", "Billable"]

    def build(self, processed_df):
        df = processed_df.copy()
        df["Task Name"] = df["Task Name Out"]
        out = df[self.COLUMNS]
        out = out[~out["Resource Name"].isin(AI_RESOURCES)]
        return out.sort_values(by=["Resource Name", "Entry Date"]).reset_index(drop=True)


class AimlOutputBuilder:
    """AI-only output. Task Name forced to 'AIML', Module sourced from input task."""

    COLUMNS = [
        "Entry Date", "Resource Name", "Task Name", "Actul Work(hrs)",
        "Teams", "Module", "Task Type", "Mandays", "Billable",
    ]
    MANDAYS_FORMULA = "={cell}/8"
    COLUMN_WIDTHS = {
        "Entry Date": 14.18,
        "Resource Name": 17.45,
        "Task Name": 10.09,
        "Actul Work(hrs)": 19.91,
        "Teams": 27.09,
        "Module": 20.45,
        "Task Type": 19.82,
        "Mandays": 18.63,
        "Billable": 11.82,
    }
    CENTERED_COLUMNS = ["Entry Date", "Task Name", "Actul Work(hrs)", "Teams", "Task Type", "Mandays"]

    def build(self, processed_df):
        ai = processed_df[processed_df["Resource Name"].isin(AI_RESOURCES)].copy()
        task_orig = ai["Task Name"].astype(str)
        out = pd.DataFrame({
            "Entry Date": ai["Entry Date"].values,
            "Resource Name": ai["Resource Name"].values,
            "Task Name": "AIML",
            "Actul Work(hrs)": ai["Actul Work(hrs)"].values,
            "Teams": "",
            "Module": [AI_MODULE_RENAME.get(t, t) for t in task_orig],
            "Task Type": [AI_ROLE_MAP.get(r, "") for r in ai["Resource Name"]],
            "Mandays": "",
            "Billable": ["No" if "leave" in t.lower() else "Yes" for t in task_orig],
        })
        return out.sort_values(by=["Resource Name", "Entry Date"]).reset_index(drop=True)


# ============================================================================
# Excel writer
# ============================================================================

class ExcelOutputWriter:
    """Writes a DataFrame to xlsx with header/body styling, column widths and per-column alignment."""

    def __init__(self, header_font=HEADER_FONT, body_font=BODY_FONT, sheet_name=SHEET_NAME):
        self.header_font = header_font
        self.body_font = body_font
        self.sheet_name = sheet_name

    def write(self, df, path, mandays_formula, column_widths=None, centered_columns=None):
        df.to_excel(path, index=False)
        wb = load_workbook(path)
        ws = wb.active
        ws.title = self.sheet_name

        header_map = {cell.value: cell.column for cell in ws[1]}
        actul_idx = header_map.get("Actul Work(hrs)")
        mandays_idx = header_map.get("Mandays")

        if actul_idx and mandays_idx:
            actul_letter = get_column_letter(actul_idx)
            for r in range(2, ws.max_row + 1):
                ws.cell(
                    row=r,
                    column=mandays_idx,
                    value=mandays_formula.format(cell=f"{actul_letter}{r}"),
                )

        for cell in ws[1]:
            cell.font = self.header_font
            cell.alignment = HEADER_ALIGNMENT

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = self.body_font

        if centered_columns:
            for col_name in centered_columns:
                idx = header_map.get(col_name)
                if not idx:
                    continue
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=idx).alignment = CENTER_ALIGNMENT

        if column_widths:
            for col_name, width in column_widths.items():
                idx = header_map.get(col_name)
                if idx:
                    ws.column_dimensions[get_column_letter(idx)].width = width

        wb.save(path)


# ============================================================================
# UI
# ============================================================================

class ExcelProcessorApp:
    BILLING_FILENAME = "Billing-Output-File.xlsx"
    AIML_FILENAME = "AIML-Output-File.xlsx"

    def __init__(self, root):
        self.root = root
        self.root.title("AI Bill - Excel Processor")
        self.file_path = None
        self.df = None

        self.processor = BillingProcessor()
        self.billing_builder = BillingOutputBuilder()
        self.aiml_builder = AimlOutputBuilder()
        self.writer = ExcelOutputWriter()

        self._build_ui()

    def _build_ui(self):
        tk.Label(self.root, text="AI Bill - Excel Processor", font=("Arial", 16, "bold")).pack(pady=10)
        self.upload_btn = tk.Button(self.root, text="Upload File", command=self.upload_file, width=20)
        self.upload_btn.pack(pady=5)
        self.process_btn = tk.Button(
            self.root, text="Process and Save File", command=self.process_and_save,
            width=20, state=tk.DISABLED,
        )
        self.process_btn.pack(pady=5)
        self.status_label = tk.Label(self.root, text="", fg="blue")
        self.status_label.pack(pady=10)

    def upload_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel File", filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if not path:
            return
        try:
            engine = "xlrd" if path.lower().endswith(".xls") else None
            self.df = pd.read_excel(path, engine=engine) if engine else pd.read_excel(path)
            self.file_path = path
            self.status_label.config(text=f"Loaded: {os.path.basename(path)}", fg="green")
            self.process_btn.config(state=tk.NORMAL)
        except Exception as e:
            self.status_label.config(text=f"Error loading file: {e}", fg="red")
            self.process_btn.config(state=tk.DISABLED)

    def process_and_save(self):
        if self.df is None:
            self.status_label.config(text="No file loaded.", fg="red")
            return
        try:
            processed = self.processor.process(self.df)
            df_billing = self.billing_builder.build(processed)
            df_aiml = self.aiml_builder.build(processed)

            folder = filedialog.askdirectory(title="Choose output folder")
            if not folder:
                self.status_label.config(text="Save cancelled.", fg="blue")
                return

            self.writer.write(
                df_billing,
                os.path.join(folder, self.BILLING_FILENAME),
                self.billing_builder.MANDAYS_FORMULA,
                column_widths=self.billing_builder.COLUMN_WIDTHS,
                centered_columns=self.billing_builder.CENTERED_COLUMNS,
            )
            self.writer.write(
                df_aiml,
                os.path.join(folder, self.AIML_FILENAME),
                self.aiml_builder.MANDAYS_FORMULA,
                column_widths=self.aiml_builder.COLUMN_WIDTHS,
                centered_columns=self.aiml_builder.CENTERED_COLUMNS,
            )
            self.status_label.config(
                text=f"Saved: {self.BILLING_FILENAME}, {self.AIML_FILENAME} in {os.path.basename(folder)}",
                fg="green",
            )
        except Exception as e:
            self.status_label.config(text=f"Error: {e}", fg="red")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()
